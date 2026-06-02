"""
生成用于测试的样例 Excel 文件。

运行方式:
    python create_sample_data.py

将在 data/excels/ 目录下生成 4 个文件：
  - 销售部_周报_2025-W23.xlsx   (正常)
  - 研发部_周报_2025-W23.xlsx   (正常)
  - 市场部_周报_2025-W23.xlsx   (正常)
  - 格式错误部门_周报_2025-W23.xlsx  (格式错误，用于异常验收)
"""

from pathlib import Path

import openpyxl


def write_dept_excel(
    path,
    dept_name: str,
    period: str,
    completed: list,
    metrics: list,       # [(name, value, mom_str), ...]
    next_week: list,
    risks: list,
):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "周报"
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14

    r = 1

    def row(a=None, b=None, c=None):
        nonlocal r
        if a is not None:
            ws.cell(r, 1, a)
        if b is not None:
            ws.cell(r, 2, b)
        if c is not None:
            ws.cell(r, 3, c)
        r += 1

    row("部门名称", dept_name)
    row("汇报周期", period)
    row()                              # blank

    row("本周完成事项")
    for item in completed:
        row(f"- {item}")
    row()

    row("本周关键指标")
    row("指标名", "数值", "环比变化")
    for name, value, mom in metrics:
        row(name, value, mom)
    row()

    row("下周计划")
    for item in next_week:
        row(f"- {item}")
    row()

    row("风险与问题")
    for item in risks:
        row(f"- {item}")

    wb.save(path)
    print(f"  Created: {path}")


def main():
    out = Path("data/excels")
    out.mkdir(parents=True, exist_ok=True)

    # ── 销售部 ─────────────────────────────────────────────────────────────
    write_dept_excel(
        out / "销售部_周报_2025-W23.xlsx",
        dept_name="销售部",
        period="2025-W23",
        completed=[
            "完成华东区 Q2 销售回款，回款率达 92%",
            "新签大客户 3 家，合同金额共计 580 万",
            "举办线上产品推介会，参与客户 120+",
        ],
        metrics=[
            ("月销售额(万元)", 1250,  "+8.3%"),
            ("新客户转化率(%)", 18.5, "-2.1%"),
            ("客户满意度(分)",  4.6,  "+0.1%"),
            ("销售漏斗转化(%)", 12.3, "-15.2%"),   # 下降超 10%，DeepSeek 应特别指出
        ],
        next_week=[
            "推进华南区 Q3 销售计划启动",
            "跟进本周未完成的潜在客户 5 家",
            "组织内部销售技能培训",
        ],
        risks=[
            "华南区竞品降价，价格压力增大",
            "销售漏斗转化率下降明显，需排查原因",
        ],
    )

    # ── 研发部 ─────────────────────────────────────────────────────────────
    write_dept_excel(
        out / "研发部_周报_2025-W23.xlsx",
        dept_name="研发部",
        period="2025-W23",
        completed=[
            "完成 v2.3 版本需求评审，确认排期",
            "修复线上 P0 Bug 2 个、P1 Bug 5 个",
            "完成用户中心模块重构，单测覆盖率达 85%",
            "发布内测版本给 QA 团队",
        ],
        metrics=[
            ("迭代完成率(%)",     87.5, "+5.0%"),
            ("Bug 修复率(%)",     92.0, "+3.0%"),
            ("代码评审通过率(%)", 96.0, "+1.0%"),
            ("系统可用性(%)",     99.95, "-0.02%"),
        ],
        next_week=[
            "启动 v2.3 版本第一阶段开发",
            "完成性能压测报告",
            "开展新人代码规范培训",
        ],
        risks=[],
    )

    # ── 市场部 ─────────────────────────────────────────────────────────────
    write_dept_excel(
        out / "市场部_周报_2025-W23.xlsx",
        dept_name="市场部",
        period="2025-W23",
        completed=[
            "发布品牌升级白皮书，行业媒体转载 12 家",
            "运营官方公众号，本周新增粉丝 2800+",
            "完成 Q2 投放效果复盘报告",
        ],
        metrics=[
            ("官网 UV(万)",      8.2,  "+12.5%"),
            ("线索转化量(条)",   156,  "-18.6%"),  # 下降超 10%
            ("内容阅读量(万)",   45.3, "+23.1%"),
            ("广告 ROI",         2.8,  "-5.0%"),
        ],
        next_week=[
            "启动 Q3 品牌推广方案执行",
            "策划 7 月行业峰会参展方案",
            "优化 SEM 投放关键词，提升转化效率",
        ],
        risks=[
            "线索转化量下降明显（-18.6%），需与销售联合分析漏斗断点",
        ],
    )

    # ── 格式错误文件（用于异常验收）────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "这是一个格式错误的文件"
    ws["A2"] = "没有按照规定模板填写，缺少部门名称和汇报周期"
    ws["A3"] = "随意内容 1"
    ws["B3"] = "随意内容 2"
    wb.save(out / "格式错误部门_周报_2025-W23.xlsx")
    print(f"  Created: {out / '格式错误部门_周报_2025-W23.xlsx'}")

    print("\n样例数据生成完成！")


if __name__ == "__main__":
    main()
